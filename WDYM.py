import os
import random
import string
import time 
import openpyxl

def randomString(stringLength=2):
    """Generate a random string of fixed length """
    letters = string.ascii_lowercase
    return ''.join(random.choice(letters) for i in range(stringLength))

l = ['o','k','ñ','.','l']
s = ['w','a','s','d','x']

r = ['4','e','r','t','f']
m = ['j','n','m',',']

loc = ("./comandos.xlsx") 

if(os.path.exists(loc)):
    wb =openpyxl.load_workbook(loc)
else:
    wb = openpyxl.Workbook()
    wb.save(loc)

comandos = ['ls','rm']

sheet = wb.active


for i in range(sheet.max_row):
    tem = sheet.cell(i+1,1).value
    if tem:
        comandos.append(tem)


max = sheet.max_row
while True:
      
    comando = input(">>>")
    if comando == '': #evitar quiebre con enter
        pass
    elif(comando == 'file'): #crear archivos random para borrar
        os.system("type NUL > " + randomString()+".txt")
    elif(comando == "ex" or comando == "exit"):#salida
        wb.save(loc)
        print("Adios :D")
        break
    else:
        #si se guardó el comando solo ejecutarlo
        if(len(comando) > 1):
            if comando[:2] in comandos:
                if(comando[0] in l or comando[0] in s):
                    os.system("dir")
                elif(comando[0] in r or comando[0] in m):
                    os.system("del"+comando[2::])
            else:
                #ls
                if(comando[0] in l):
                    if(comando[1] in s):
                        res = input("Did you mean ls ? (Y/N)")
                        if(res == 'Y' or res == 'y'):
                            comandos.append(comando[:2])   
                            max = max + 1    
                            sheet.cell(max,1).value = comando[:2]      
                            os.system("dir")
                    else:
                        print("Este comando no se reconoce")
                elif(comando[0] in s):
                    if(comando[1] in l):
                        res = input("Did you mean ls ? (Y/N)")
                        if(res == 'Y' or res == 'y'):
                            comandos.append(comando[:2])  
                            max = max + 1    
                            sheet.cell(max,1).value = comando[:2]                      
                            os.system("dir")
                    else:
                        print("Este comando no se reconoce")

                #rm
                elif(comando[0] in r):
                    if(comando[1] in m):
                        res = input("Did you mean rm ? (Y/N)")
                        if(res == 'Y' or res == 'y'):
                            comandos.append(comando[:2])  
                            max = max + 1    
                            sheet.cell(max,1).value = comando[:2]                      
                            os.system("del"+comando[2::])
                    else:
                        print("Este comando no se reconoce")
                elif(comando[0] in m):
                    if(comando[1] in r):
                        res = input("Did you mean rm ? (Y/N)")
                        if(res == 'Y' or res == 'y'):
                            comandos.append(comando[:2]) 
                            max = max + 1    
                            sheet.cell(max,1).value = comando[:2]                       
                            os.system("del"+comando[2::])
                    else:
                        print("Este comando no se reconoce")
                
        else:
            print("Este comando no se reconoce")